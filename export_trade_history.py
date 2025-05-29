import MetaTrader5 as mt5
import openpyxl
from openpyxl.utils import coordinate_to_tuple
from openpyxl.utils.cell import coordinate_from_string
from datetime import datetime
from copy import copy

# 定数を定義
TIME_SHIFT = 60 * 60 * 9                                      # 時刻を9時間シフトするための定数
FILE_PATH = fr"C:\Users\username\tradenote_template.xlsx"     # トレードノートのパス
SHEET_NAME = "トレード履歴"                                    # シート名
TABLE_NAME = "trade_history"                                  # テーブル名
ROW_HEIGHT = 75                                               # 行の高さ

# 関数を定義
def return_shifted_unix_time_from_string(date_string):
    if "T" in date_string:
        try:
            unix_time = datetime.strptime(date_string, "%Y/%m/%dT%H:%M:%S")
        except ValueError:
            return None
        
        unix_time = datetime.timestamp(unix_time)
    elif "t" in date_string:
        try:
            unix_time = datetime.strptime(date_string, "%Y/%m/%dt%H:%M:%S")
        except ValueError:
            return None
        
        unix_time = datetime.timestamp(unix_time)
    else:
        try:
            unix_time = datetime.strptime(date_string, "%Y/%m/%d")
        except ValueError:
            return None
        
        unix_time = datetime.timestamp(unix_time)
    return unix_time

def change_weekday_to_japanese(weekday):
    match weekday:
        case 0:
            return "月"
        case 1:
            return "火"
        case 2:
            return "水"
        case 3:
            return "木"
        case 4:
            return "金"
        case 5:
            return "土"
        case 6:
            return "日"
        case _:
            return ""
        
def change_type_to_string(type):
    match type:
        case mt5.DEAL_TYPE_BUY:
            return "buy"
        case mt5.DEAL_TYPE_SELL:
            return "sell"
        case mt5.DEAL_TYPE_BALANCE:
            return "balance"
        case _:
            return "else"

def round_price(price):
    if price < 10:
        return round(price,5)
    elif price < 1000:
        return round(price,3)
    else:
        return round(price,2)

def copy_cell_style(dst_cell, src_cell):
    dst_cell.font = copy(src_cell.font)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.border = copy(src_cell.border)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.number_format = copy(src_cell.number_format)
    return

def expand_table_range_by_one_row(table):
    ref = table.ref
    start_cell, end_cell = ref.split(":")
    # 行・列番号に変換
    end_col_letter, end_row = coordinate_from_string(end_cell)
    table.ref = f"{start_cell}:{end_col_letter}{end_row+1}"

def ask_yes_no(prompt="続けますか？ [y/n]: "):
    while True:
        answer = input(prompt).strip().lower()
        if answer in ("y", "yes"):
            return True
        elif answer in ("n", "no"):
            return False
        else:
            print("y または n で答えてください。")
            ask_yes_no(prompt)

def main():
    # MT5への接続を初期化
    if not mt5.initialize():
        print("MT5の接続に失敗しました")
        mt5.shutdown()
        exit()

    while(True):
        # エクセルのワークブック、ワークシート、テーブルを取得
        try:
            wb = openpyxl.load_workbook(FILE_PATH)
        except FileNotFoundError:
            print("ファイルパスが不適切です")
            break
        try:
            ws = wb[SHEET_NAME]
        except KeyError:
            print("ワークシート名が不適切です")
            break
        try:
            table = ws.tables[TABLE_NAME]
        except KeyError:
            print("テーブル名が不適切です")
            break

        # input
        from_date_input = input("Enter from_date(e.g. 2024/1/14T14:30:00 or 2024/1/14): ")
        from_date = return_shifted_unix_time_from_string(from_date_input)
        if(from_date == None):
            print("入力がフォーマットに一致しません")
            if(ask_yes_no("トレード履歴の出力を続けますか？[y/n]: ")):
                continue
            else:
                break
        to_date_input = input("Enter to_date(e.g. 2024/1/14T22:20:00 or 2024/1/14): ")
        to_date = return_shifted_unix_time_from_string(to_date_input)
        if(to_date == None):
            print("入力がフォーマットに一致しません")
            if(ask_yes_no("トレード履歴の出力を続けますか？[y/n]: ")):
                continue
            else:
                break
            
        # 取得期間のエラー処理
        if from_date > to_date:
            print("Error: from_dateがto_dateよりも後になっています")
            if(ask_yes_no("トレード履歴の出力を続けますか？[y/n]: ")):
                continue
            else:
                break

        # 時刻をTIME_SHIFTだけずらす
        from_date += TIME_SHIFT
        to_date += TIME_SHIFT
        
        symbol_input = input("Enter symbol(e.g. usdjpy): ")
        group_input = "*" + symbol_input.upper() + "*"

        # 取引履歴を取得
        deal_history = mt5.history_deals_get(from_date, to_date, group=group_input)

        # 取引履歴が取得できているかチェック
        if deal_history is None:
            print("履歴の取得に失敗しました")
            if(ask_yes_no("トレード履歴の出力を続けますか？[y/n]: ")):
                continue
            else:
                break
        elif len(deal_history) == 0:
            print("該当するデータがありませんでした")
            if(ask_yes_no("トレード履歴の出力を続けますか？[y/n]: ")):
                continue
            else:
                break

        # テーブルの最終行の1行下の行に出力するようにする
        row_writing, _ = coordinate_to_tuple(table.ref.split(":")[1])
        row_writing += 1

        # テーブル範囲を変更
        expand_table_range_by_one_row(table)
        # 行の高さを変更
        ws.row_dimensions[row_writing].height = ROW_HEIGHT

        trade_date = ""
        trade_symbol = ""
        trade_type = 0
        trade_entry_price  = 0
        trade_stoploss_price = 0
        trade_close_price = 0
        trade_entry_volume = 0
        trade_close_volume = 0
        trade_profit = 0
        trade_commission = 0
        trade_swap = 0
        trade_total_profit = 0
        trade_riskreward = 0

        # トレード履歴をワークシートに追加
        for idx, deal in enumerate(deal_history):
            # 1個目の取引からトレード日、通貨ペア、エントリー方向を決定する
            if idx == 0:
                # TIME_SHIFTの分だけ時間をずらす
                tmp_date = datetime.fromtimestamp(deal.time - TIME_SHIFT)
                trade_date = tmp_date.strftime("%Y/%m/%d")
                trade_day_of_the_week = change_weekday_to_japanese(tmp_date.weekday())
                trade_symbol = deal.symbol
                trade_type = deal.type

            # 通貨ペアが異なる場合
            if deal.symbol != trade_symbol:
                print("Error: 複数の通貨ペアのトレード履歴が含まれます")
                if(ask_yes_no("トレード履歴の出力を続けますか？[y/n]: ")):
                    continue
                else:
                    break

            # エントリー/決済 共通
            trade_profit += deal.profit
            trade_commission += deal.commission + deal.fee
            trade_swap += deal.swap
            
            # エントリーの場合
            if (deal.entry == mt5.DEAL_ENTRY_IN and deal.type == trade_type):
                trade_entry_volume += deal.volume
                trade_entry_price += deal.price*deal.volume
                # 注文履歴からsl価格を取得
                order = mt5.history_orders_get(ticket=deal.order)
                trade_stoploss_price += order[0].sl*deal.volume
            # 決済の場合
            elif (deal.entry == mt5.DEAL_ENTRY_OUT and deal.type == 1-trade_type):
                trade_close_volume += deal.volume
                trade_close_price += deal.price*deal.volume
            
        if trade_entry_volume != trade_close_volume:
            print("Error: エントリーと決済のロット数が合致しません")
            if(ask_yes_no("トレード履歴の出力を続けますか？[y/n]: ")):
                continue
            else:
                break

        trade_type = change_type_to_string(trade_type)
        trade_entry_price = round_price(trade_entry_price/trade_entry_volume)
        trade_stoploss_price = round_price(trade_stoploss_price/trade_entry_volume)
        trade_close_price = round_price(trade_close_price/trade_entry_volume)
        trade_total_profit = trade_profit + trade_commission + trade_swap
        trade_riskreward = (trade_close_price - trade_entry_price)/(trade_entry_price - trade_stoploss_price)

        row = [
            trade_date,                   # 日付
            trade_day_of_the_week,        # 曜日
            trade_symbol,                 # シンボル
            trade_type,                   # エントリー方向
            trade_entry_price,            # エントリー価格
            trade_stoploss_price,         # ストップロス価格
            trade_close_price,            # 決済価格
            trade_entry_volume,           # ロット数
            trade_profit,                 # 利益
            trade_commission,             # 手数料
            trade_swap,                   # スワップ
            trade_total_profit,           # 総利益
            trade_riskreward              # リスクリワード
        ]

        # データを出力
        for col_writing, value in enumerate(row, start=1):
            cell_writing = ws.cell(row=row_writing, column=col_writing)
            cell_above = ws.cell(row=row_writing-1, column=col_writing)
            copy_cell_style(cell_writing, cell_above)

            ws.cell(row=row_writing, column=col_writing, value=value)

        # Excelファイルを保存
        wb.save(FILE_PATH)
        print(f"トレード履歴が {FILE_PATH} に保存されました")
        if(ask_yes_no("トレード履歴の出力を続けますか？[y/n]: ")):
            continue
        else:
            break

    print("トレード履歴の出力を終了します")
    # MT5の終了
    mt5.shutdown()

if __name__ == "__main__":
    main()










